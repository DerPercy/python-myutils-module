import os
import logging
import glob
import re
import copy
from ..valueParser import getValue

import openpyxl


class XLSXStorage:
    settings = {}
    def __init__(self,settings):
        self.settings = settings
        pass
    def query(self,query):
        returnList = []
        class Methods:
            pass

        methods = Methods()
        obj_glob = {}
        obj_regex = {}
        for elm in self.settings["xlsx"]["properties"]:
            obj_glob[elm] = "*"
            obj_regex[elm] = self.settings["xlsx"]["properties"][elm]["pattern"]

        fname_original = self.settings["xlsx"]["filepattern"]
        logging.debug("Filename before replacement: "+fname_original)
        fname_glob = getValue(fname_original,obj_glob,methods)
        logging.debug("glob selector: "+fname_glob)
        fname_regex = fname_original

        fname_regex = regex_prepare(fname_regex)
        fname_regex = getValue(fname_regex,obj_regex,methods)

        logging.debug("regex selector: "+fname_regex)

        filenamesList = glob.glob(fname_glob)
        for fname in filenamesList:
            logging.debug("Checking file:"+fname)
            z = re.match(fname_regex,fname)
            if z:
                logging.debug("Filename valid")
                fileobj = {}
                extractValues(fname,fname_original,self.settings["xlsx"]["properties"],fileobj)
                returnList.extend(readXLSX(fname,self.settings["xlsx"]["filecontent"],fileobj))
                #readXLSX(fname,self.settings["xlsx"]["filecontent"],fileobj)
                #logging.debug(z.groups()[0])
            else:
                logging.debug("Filename invalid.Skipping")

        return returnList

def readXLSX(filename,settings_filecontent,base_obj):
    retList = []
    logging.debug("Reading"+filename)
    wb_obj = openpyxl.load_workbook(filename)
    sheet = wb_obj.active
    col_names = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i >= settings_filecontent["startrow"]:
            # Check for Stopping at empty values
            if "stopatnonevalue" in settings_filecontent:
                if row[int(settings_filecontent["stopatnonevalue"])] == None:
                    return retList
            #append
            sub_obj = copy.deepcopy(base_obj)
            #print(i)
            #print(row)
            for col in settings_filecontent["columns"]:
                attrName = settings_filecontent["columns"][col];
                #print(attrName)
                #print(col)
                sub_obj[attrName] = row[int(col)]
            retList.append(sub_obj)
        #for column in sheet.iter_cols(i, sheet.max_column):
            #col_names.append(column[0].value)
            #print(colums)
    #print(col_names)
    return retList

def extractValues(content,pattern,properties,object):
    """
    Extract the values from the content based on the pattern and properties
    and store them into the object
    """
    class Methods:
        pass
    methods = Methods()
    obj_regex = {}
    for elm in properties:
        obj_regex[elm] = properties[elm]["pattern"]
    cont_pattern = getValue(pattern,obj_regex,methods)
    logging.debug(cont_pattern)
    for elm in properties:
        obj_group = copy.deepcopy(obj_regex)
        obj_group[elm] = "("+properties[elm]["pattern"]+")"
        re_pattern = regex_prepare(pattern)
        cont_group = getValue(re_pattern,obj_group,methods)
        z = re.match(cont_group,content)
        if z:
            if len(z.groups()) > 0:
                object[elm] = z.groups()[0]
    pass

def regex_prepare(content):
    """
    Prepare a string to be a valid regex pattern
    """
    content = content.replace("\\","\\\\")
    content = content.replace("/","\\/")
    content = content.replace(".","\\.")
    return content
